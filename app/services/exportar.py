"""Excel export for grading results using openpyxl."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.db.models.alumno import Alumno
from app.db.models.entrega import Entrega
from app.db.models.revision import Revision
from app.db.models.tarea import Tarea

_HEADER_FILL = PatternFill(start_color="1B4F8A", end_color="1B4F8A", fill_type="solid")
_ALT_FILL = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def generar_excel(tarea_id: int, db: Session) -> bytes:
    tarea: Tarea = db.get(Tarea, tarea_id)
    criterios = tarea.criterios or []

    entregas = (
        db.query(Entrega)
        .join(Alumno)
        .filter(Entrega.tarea_id == tarea_id)
        .order_by(Alumno.nombre)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Build header row
    headers = ["Alumno", "Calificación"]
    for c in criterios:
        peso = int(float(c.get("peso", 0)) * 100)
        headers.append(f"{c.get('nombre', '')} ({peso}%)")
    headers += ["Retroalimentación", "% IA", "Nivel IA", "Fragmentos sospechosos"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, entrega in enumerate(entregas, start=2):
        revision: Revision | None = entrega.revision
        fill = _ALT_FILL if row_idx % 2 == 0 else PatternFill()

        # Build desglose map keyed by criterio name
        desglose_map: dict[str, float] = {}
        if revision and revision.desglose:
            for item in revision.desglose:
                desglose_map[item.get("criterio", "")] = item.get("puntos", 0)

        # Suspicious fragments as readable string
        fragmentos_str = ""
        if revision and revision.ia_fragmentos:
            fragmentos_str = " | ".join(
                f['texto'][:80] for f in revision.ia_fragmentos if 'texto' in f
            )

        row_data = [entrega.alumno.nombre, revision.calificacion if revision else None]
        for c in criterios:
            row_data.append(desglose_map.get(c.get("nombre", ""), None))
        row_data.append(revision.retroalimentacion if revision else None)
        row_data.append((revision.ia_probabilidad / 100) if revision and revision.ia_probabilidad is not None else None)
        row_data.append(revision.ia_nivel_riesgo if revision else None)
        row_data.append(fragmentos_str)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Calificación: 1 decimal
        ws.cell(row=row_idx, column=2).number_format = "0.0"
        # % IA column index
        ia_col = 2 + len(criterios) + 2  # Alumno + Cal + criterios + Retro + % IA
        ws.cell(row=row_idx, column=ia_col).number_format = '0.0"%"'

    # Auto-fit column widths (max 60)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(entregas) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 60))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
